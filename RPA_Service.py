import pyautogui
import os
import random
import tkinter as tk
from tkinter import messagebox, ttk
from openpyxl import Workbook, load_workbook
from datetime import datetime

# Ruta al archivo de asistentes
RUTA_ASISTENTES = 'asistentes.xlsx'
# Ruta para el archivo donde se guardarán las consultas agendadas
RUTA_CONSULTAS = 'consultas_agendadas.xlsx'

# Función para cargar el archivo de asistentes y seleccionar uno al azar
def seleccionar_asistente_aleatorio():
    if not os.path.exists(RUTA_ASISTENTES):
        pyautogui.alert("No se encontró el archivo de asistentes.", "Error")
        return None
    
    wb = load_workbook(RUTA_ASISTENTES)
    ws = wb.active  # Establecer la hoja activa
    
    asistentes = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # Ignorar el encabezado
        if row[0]:  # Verificar que la celda tenga un valor
            asistentes.append(row[0])
    
    wb.close()
    
    if not asistentes:
        pyautogui.alert("No hay asistentes disponibles en el archivo de asistentes.", "Advertencia")
        return None

    return random.choice(asistentes)

# Función para obtener datos del usuario
def obtener_datos():
    nombre = pyautogui.prompt("Por favor, ingrese su nombre para agendar la consulta")
    if not nombre:
        pyautogui.alert("Debe ingresar un nombre válido", "Advertencia")
        return None, None, None, None

    motivo = pyautogui.prompt("Por favor, ingrese el motivo de la consulta")
    if not motivo:
        pyautogui.alert("Debe ingresar un motivo válido", "Advertencia")
        return None, None, None, None

    departamento = pyautogui.prompt("Por favor, ingrese el departamento")
    if not departamento:
        pyautogui.alert("Debe ingresar un departamento válido", "Advertencia")
        return None, None, None, None

    asistente = seleccionar_asistente_aleatorio()
    if not asistente:
        pyautogui.alert("No hay asistentes disponibles para asignar.", "Advertencia")
        return None, None, None, None

    return nombre, motivo, departamento, asistente

# Función para crear o cargar el archivo Excel para guardar las consultas
def obtener_archivo_excel():
    if not os.path.exists(RUTA_CONSULTAS):
        wb = Workbook()
        ws = wb.active
        ws.title = "Consultas Agendadas"
        ws.append(["Nombre", "Motivo", "Departamento", "Asistente", "Fecha y Hora", "Estado"])
        wb.save(RUTA_CONSULTAS)
    return RUTA_CONSULTAS

# Función para agendar la consulta y guardarla en el archivo de Excel
def agendar_consulta(nombre, motivo, departamento, asistente):
    ruta_archivo = obtener_archivo_excel()
    wb = load_workbook(ruta_archivo)
    ws = wb["Consultas Agendadas"]
    fecha_hora = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws.append([nombre, motivo, departamento, asistente, fecha_hora, "Abierto"])
    wb.save(ruta_archivo)
    pyautogui.alert(f"Consulta agendada con éxito para {nombre}.", "Éxito")

# Interfaz gráfica para gestionar los tickets
class TicketApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Gestión de Consultas")
        
        # Configurar la ventana para pantalla completa
        self.root.attributes("-fullscreen", True)
        
        self.tree = ttk.Treeview(root, columns=("Nombre", "Motivo", "Departamento", "Asistente", "Fecha y Hora", "Estado"), show='headings')
        
        for col in self.tree["columns"]:
            self.tree.heading(col, text=col)
        
        self.tree.pack(fill="both", expand=True)
        
        # Botones para abrir/cerrar tickets
        abrir_btn = tk.Button(root, text="Abrir Ticket", command=self.abrir_ticket)
        cerrar_btn = tk.Button(root, text="Cerrar Ticket", command=self.cerrar_ticket)
        abrir_btn.pack(side="left", padx=10, pady=10)
        cerrar_btn.pack(side="left", padx=10, pady=10)
        
        self.cargar_tickets()

    # Cargar tickets en la interfaz desde el archivo Excel
    def cargar_tickets(self):
        ruta_archivo = obtener_archivo_excel()
        wb = load_workbook(ruta_archivo)
        ws = wb["Consultas Agendadas"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            self.tree.insert("", "end", values=row)
    
    # Función para marcar un ticket como abierto
    def abrir_ticket(self):
        self.actualizar_estado("Abierto")
    
    # Función para marcar un ticket como cerrado
    def cerrar_ticket(self):
        self.actualizar_estado("Cerrado")

    # Actualizar estado del ticket seleccionado
    def actualizar_estado(self, nuevo_estado):
        seleccion = self.tree.selection()
        if not seleccion:
            messagebox.showwarning("Advertencia", "Seleccione un ticket para actualizar.")
            return
        
        ruta_archivo = obtener_archivo_excel()
        wb = load_workbook(ruta_archivo)
        ws = wb["Consultas Agendadas"]
        
        for item in seleccion:
            valores = self.tree.item(item, "values")
            for row in ws.iter_rows(min_row=2):
                if row[0].value == valores[0] and row[4].value == valores[4]:  # Verificación por Nombre y Fecha
                    row[5].value = nuevo_estado
                    self.tree.item(item, values=(valores[0], valores[1], valores[2], valores[3], valores[4], nuevo_estado))
                    break
        
        wb.save(ruta_archivo)
        messagebox.showinfo("Actualización", f"El ticket ha sido marcado como {nuevo_estado}.")

# Ejecutar la función principal
def main():
    nombre, motivo, departamento, asistente = obtener_datos()
    if nombre and motivo and departamento and asistente:
        agendar_consulta(nombre, motivo, departamento, asistente)
        root = tk.Tk()
        app = TicketApp(root)
        root.mainloop()

if __name__ == "__main__":
    main()